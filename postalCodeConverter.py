# import pandas lib as pd
import json

import openpyxl
import pandas as pd

# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel('postalCodeordered.xlsx')

dataframe = openpyxl.load_workbook('postalCodeordered.xlsx')

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
postalCodeDb = {}
for row in range(0, dataframe1.max_row):
    postalCodeDict = dict()
    postalCoedSet = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        postalCoedSet.append(col[row].value)
        # print(col[row].value)
    if (postalCoedSet[0] not in [None, '\xa0']):
        postalCodeDict["division"] = postalCoedSet[0]
        postalCodeDict["district"] = postalCoedSet[1]
        postalCodeDict["postalCode"] = postalCoedSet[2]

    postalCodeDb[postalCoedSet[2]] = postalCodeDict

    # print(len(postalCodeDict))


jsonDumpDb = postalCodeDb


with open('postalCodeDB.json', 'w') as f:
    json.dump(jsonDumpDb, f)


print(jsonDumpDb)
